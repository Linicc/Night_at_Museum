"""Write simulation results to Excel file."""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from core.ca.ca_grid import (
    CELL_EMPTY, CELL_PERSON, CELL_WALL, CELL_EXIT,
    CELL_ENTRANCE, CELL_EXHIBIT, CELL_EXHIBIT_SPECIAL, CELL_SECURITY
)


class ExcelWriter:
    """Write CA simulation results to Excel workbook."""

    def __init__(self, width=100, height=100):
        """Initialize Excel writer."""
        self.width = width
        self.height = height
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.step_counter = 0

    def add_config_sheet(self, grid_data):
        """Add initial config sheet to workbook."""
        ws = self.wb.create_sheet("Config", 0)

        for x in range(self.width):
            for y in range(self.height):
                cell_value = grid_data[x][y] if x < len(grid_data) and y < len(grid_data[x]) else 0
                ws.cell(row=y + 1, column=x + 1).value = cell_value

    def add_timestep_snapshot(self, timestep, grid_snapshot, skip_interval=100):
        """Add timestep snapshot sheet (every skip_interval steps).

        Args:
            timestep: Current simulation timestep
            grid_snapshot: 2D grid of cell values
            skip_interval: Only add sheet for timesteps divisible by this
        """
        if timestep % skip_interval != 0:
            return

        sheet_name = f"Timestep_{timestep:04d}"
        ws = self.wb.create_sheet(sheet_name)

        for x in range(self.width):
            for y in range(self.height):
                cell_value = grid_snapshot[x][y] if x < len(grid_snapshot) and y < len(grid_snapshot[x]) else 0
                cell = ws.cell(row=y + 1, column=x + 1)
                cell.value = cell_value

                # Color cells by type
                self._color_cell(cell, cell_value)

    def add_agent_trajectories(self, logger_data):
        """Add agent trajectories sheet."""
        ws = self.wb.create_sheet("AgentTrajectories")

        # Headers
        ws['A1'] = "Timestep"
        ws['B1'] = "AgentID"
        ws['C1'] = "X"
        ws['D1'] = "Y"
        ws['E1'] = "PanicLevel"
        ws['F1'] = "Evacuated"
        ws['G1'] = "Age"

        row = 2
        for record in logger_data:
            ws.cell(row=row, column=1).value = record['timestep']
            ws.cell(row=row, column=2).value = record['agent_id']
            ws.cell(row=row, column=3).value = record['x']
            ws.cell(row=row, column=4).value = record['y']
            ws.cell(row=row, column=5).value = round(record['panic_level'], 3)
            ws.cell(row=row, column=6).value = record['evacuated']
            ws.cell(row=row, column=7).value = record['age']
            row += 1

        # Auto-width columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 12

    def add_summary_sheet(self, simulation_stats):
        """Add summary statistics sheet."""
        ws = self.wb.create_sheet("Summary")

        ws['A1'] = "Metric"
        ws['B1'] = "Value"

        metrics = [
            ("Total Timesteps", simulation_stats.get('total_timesteps', 0)),
            ("Total Agents", simulation_stats.get('total_agents', 0)),
            ("Evacuated Agents", simulation_stats.get('evacuated_agents', 0)),
            ("Evacuation Rate (%)",
             (simulation_stats.get('evacuated_agents', 0) / max(1, simulation_stats.get('total_agents', 1))) * 100),
            ("Average Panic Level", round(simulation_stats.get('avg_panic_final', 0.0), 3)),
            ("Max Panic Level", round(simulation_stats.get('max_panic_final', 0.0), 3)),
            ("Average Stamina Final", round(simulation_stats.get('avg_stamina_final', 1.0), 3)),
        ]

        row = 2
        for metric_name, value in metrics:
            ws.cell(row=row, column=1).value = metric_name
            ws.cell(row=row, column=2).value = value
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def save(self, filepath):
        """Save workbook to file."""
        os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
        self.wb.save(filepath)
        self.wb.close()

    def _color_cell(self, cell, cell_type):
        """Apply color to cell based on type."""
        colors = {
            CELL_EMPTY: "FFFFFF",        # White
            CELL_PERSON: "0070C0",       # Blue
            CELL_WALL: "000000",         # Black
            CELL_EXIT: "70AD47",         # Green
            CELL_ENTRANCE: "FFC000",     # Yellow
            CELL_EXHIBIT: "FF7030",      # Orange
            CELL_EXHIBIT_SPECIAL: "FF00FF",  # Magenta
            CELL_SECURITY: "C55A11",     # Brown
        }

        color = colors.get(cell_type, "FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Adjust text color for dark backgrounds
        if cell_type in [CELL_WALL, CELL_EXHIBIT_SPECIAL]:
            cell.font = Font(color="FFFFFF")


def create_output_workbook(simulation, logger_data, output_path):
    """Create complete output workbook with all sheets.

    Args:
        simulation: CASimulation instance
        logger_data: List of agent trajectory records from logger
        output_path: Path to save Excel file
    """
    writer = ExcelWriter(simulation.width, simulation.height)

    # Add config sheet
    writer.add_config_sheet(simulation.grid.static_layer)

    # Add agent trajectories
    writer.add_agent_trajectories(logger_data)

    # Add summary statistics
    stats = {
        'total_timesteps': simulation.timestep,
        'total_agents': len(simulation.agents),
        'evacuated_agents': len(simulation.evacuated_agents),
        'avg_panic_final': sum(a.panic_level for a in simulation.agents) / len(simulation.agents) if simulation.agents else 0.0,
        'max_panic_final': max((a.panic_level for a in simulation.agents), default=0.0),
        'avg_stamina_final': sum(a.stamina for a in simulation.agents) / len(simulation.agents) if simulation.agents else 1.0,
    }
    writer.add_summary_sheet(stats)

    # Save
    writer.save(output_path)

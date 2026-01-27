"""Parse museum configuration from Excel file."""
import os
from openpyxl import load_workbook
from core.ca.ca_grid import (
    CELL_EMPTY, CELL_PERSON, CELL_WALL, CELL_EXIT,
    CELL_ENTRANCE, CELL_EXHIBIT, CELL_EXHIBIT_SPECIAL, CELL_SECURITY
)


def parse_excel_config(filepath):
    """Parse Excel configuration file and return grid, agent positions, and parameters.

    Expected Excel structure:
    - Sheet "Config": 100×100 grid with cell type values (0-7)
    - Parameters in rows 102-110
    - Optional sheet "InitialState" for manual agent placement

    Returns:
        {
            'grid_data': 2D list of cell types,
            'width': 100,
            'height': 100,
            'agents': list of (x, y, age, family_id),
            'params': {'simulation_steps': int, 'initial_population': int, ...}
        }
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Config file not found: {filepath}")

    wb = load_workbook(filepath)

    if 'Config' not in wb.sheetnames:
        raise ValueError("Excel file must have 'Config' sheet")

    ws = wb['Config']

    # Parse grid (100×100)
    width, height = 100, 100
    grid_data = []

    for x in range(1, width + 1):
        column = []
        for y in range(1, height + 1):
            cell = ws.cell(row=y, column=x)
            value = cell.value if cell.value is not None else 0
            try:
                column.append(int(value))
            except (ValueError, TypeError):
                column.append(0)  # Default to empty
        grid_data.append(column)

    # Parse parameters from rows 102-110
    params = {
        'simulation_steps': int(ws.cell(row=102, column=1).value or 1000),
        'initial_population': int(ws.cell(row=103, column=1).value or 75),
        'panic_spread_rate': float(ws.cell(row=104, column=1).value or 0.05),
        'panic_decay_rate': float(ws.cell(row=105, column=1).value or 0.01),
        'crowding_threshold': int(ws.cell(row=106, column=1).value or 5),
    }

    # Parse agent initial positions from "InitialState" sheet if present
    agents = []
    if 'InitialState' in wb.sheetnames:
        ws_init = wb['InitialState']
        for row in range(2, ws_init.max_row + 1):  # Skip header
            agent_id = ws_init.cell(row=row, column=1).value
            x = ws_init.cell(row=row, column=2).value
            y = ws_init.cell(row=row, column=3).value
            age = ws_init.cell(row=row, column=4).value
            family_id = ws_init.cell(row=row, column=5).value

            if agent_id and x and y:
                agents.append({
                    'id': int(agent_id),
                    'x': int(x),
                    'y': int(y),
                    'age': int(age) if age else None,
                    'family_id': int(family_id) if family_id else None,
                })

    wb.close()

    return {
        'grid_data': grid_data,
        'width': width,
        'height': height,
        'agents': agents,
        'params': params,
    }


def create_empty_config_template(filepath, width=100, height=100):
    """Create an empty Excel template with 100×100 grid."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Config"

    # Initialize grid with zeros (empty cells)
    for x in range(1, width + 1):
        for y in range(1, height + 1):
            ws.cell(row=y, column=x).value = 0

    # Add parameter labels and default values
    ws.cell(row=102, column=1).value = 1000  # simulation_steps
    ws.cell(row=103, column=1).value = 75    # initial_population
    ws.cell(row=104, column=1).value = 0.05  # panic_spread_rate
    ws.cell(row=105, column=1).value = 0.01  # panic_decay_rate
    ws.cell(row=106, column=1).value = 5     # crowding_threshold

    # Add parameter labels in column B
    ws.cell(row=102, column=2).value = "simulation_steps"
    ws.cell(row=103, column=2).value = "initial_population"
    ws.cell(row=104, column=2).value = "panic_spread_rate"
    ws.cell(row=105, column=2).value = "panic_decay_rate"
    ws.cell(row=106, column=2).value = "crowding_threshold"

    # Create InitialState sheet
    ws_init = wb.create_sheet("InitialState")
    ws_init['A1'] = "AgentID"
    ws_init['B1'] = "X"
    ws_init['C1'] = "Y"
    ws_init['D1'] = "Age"
    ws_init['E1'] = "FamilyID"

    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Metric"
    ws_summary['B1'] = "Value"
    ws_summary['A2'] = "Total Timesteps"
    ws_summary['A3'] = "Evacuated Agents"
    ws_summary['A4'] = "Evacuation Time"
    ws_summary['A5'] = "Avg Panic Level"

    wb.save(filepath)
